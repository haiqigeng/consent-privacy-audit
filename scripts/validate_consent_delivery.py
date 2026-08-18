from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

from jsonschema import Draft202012Validator
from openpyxl import load_workbook

from consent_runtime_core import (
    ContractError,
    FINGERPRINT_ALGO_VERSION,
    IDENTITY_NORMALIZATION_VERSION,
    PRIORITY_RUBRIC_VERSION,
    PrivacyError,
    SCHEMA_DIR,
    assert_privacy_safe,
    finding_fingerprint,
    iter_jsonl,
    load_adapters,
    load_profile,
    path_within,
    read_json,
    sha256_file,
    technical_priority,
    validate_schema,
)
from review_screenshot_evidence import extract_text, review_text
from scan_consent_runtime import load_restricted_canaries


REQUIRED_FILES = {
    "audit-run.json",
    "observations.jsonl",
    "findings.json",
    "declaration-diff.json",
    "evidence-index.json",
    "consent-privacy-audit.md",
    "consent-privacy-audit.xlsx",
    "remediation-handoff.json",
    "monitoring-baseline.json",
}
POSITIVE_LEGAL_CLAIM_RE = re.compile(
    r"\b(?:is|are|remains?|has been|fully|therefore)\s+(?:gdpr\s+|eprivacy\s+|cnil\s+)?compliant\b|"
    r"\bcomplies\s+with\s+(?:the\s+)?(?:gdpr|eprivacy|cnil)\b|\bcertif(?:y|ies|ied)\s+(?:legal\s+)?compliance\b",
    re.I,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate structural, privacy, evidence, source, finding, and language gates.")
    parser.add_argument("delivery", type=Path, nargs="?")
    parser.add_argument("--package-only", action="store_true")
    parser.add_argument("--canary-file", type=Path, help="Restricted input used only for independent non-retention scanning")
    return parser.parse_args()


def validate_package() -> list[str]:
    checks: list[str] = []
    for path in sorted(SCHEMA_DIR.glob("*.json")):
        schema = read_json(path)
        Draft202012Validator.check_schema(schema)
        checks.append(f"schema:{path.name}")
    adapters = load_adapters()
    if {item["adapter_id"] for item in adapters} != {"axeptio-web", "didomi-web", "onetrust-web", "generic-tcf-web", "generic-custom-banner"}:
        raise ContractError("The v1 CMP adapter set is incomplete or unexpectedly expanded")
    checks.extend(f"adapter:{item['adapter_id']}" for item in adapters)
    for profile_id in ("neutral-technical", "cnil-fr"):
        profile = load_profile(profile_id)
        source_ids = {str(item["source_id"]) for item in profile["sources"]}
        for rule in profile["rules"]:
            missing = set(rule["source_ids"]) - source_ids
            if missing:
                raise ContractError(f"Profile {profile_id} rule {rule['rule_id']} references missing sources: {sorted(missing)}")
        checks.append(f"profile:{profile_id}")
    return checks


def load_findings(path: Path) -> list[dict[str, Any]]:
    value = read_json(path)
    if not isinstance(value, list):
        raise ContractError("findings.json must be an array")
    for item in value:
        validate_schema(item, "finding.schema.json", label=item.get("finding_fingerprint", "finding"))
    return value


def scan_xlsx(path: Path, canaries: list[str]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        expected = {"Summary", "Scope & Coverage", "Scenario Results", "Findings", "Declaration Diff", "Owner Backlog", "Evidence Index", "Limitations & Deferred Work"}
        if set(workbook.sheetnames) != expected:
            raise ContractError(f"Workbook sheet contract mismatch: {workbook.sheetnames}")
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        raise ContractError(f"Workbook formula is prohibited at {sheet.title}!{cell.coordinate}")
                    if isinstance(cell.value, str):
                        assert_privacy_safe(cell.value, canaries=canaries, label=f"workbook {sheet.title}!{cell.coordinate}")
    finally:
        workbook.close()


def scan_delivery_files(directory: Path, canaries: list[str], evidence: dict[str, Any]) -> None:
    forbidden_suffixes = {".har", ".netlog"}
    forbidden_names = {"restricted-canaries.json", "canaries.json", "authorization.json", "submission-authorizations.json"}
    for path in directory.rglob("*"):
        if not path.is_file():
            continue
        if path.suffix.casefold() in forbidden_suffixes or path.name.casefold() in forbidden_names:
            raise PrivacyError(f"Restricted/raw artifact is present in delivery: {path.relative_to(directory)}")
        if path.suffix.casefold() in {".json", ".jsonl", ".md", ".txt", ".csv", ".xml", ".html"}:
            text = path.read_text(encoding="utf-8-sig")
            assert_privacy_safe(text, canaries=canaries, label=str(path.relative_to(directory)))
            if path.suffix.casefold() == ".md" and POSITIVE_LEGAL_CLAIM_RE.search(text):
                raise ContractError(f"Positive legal-compliance claim found in {path.relative_to(directory)}")
    scan_xlsx(directory / "consent-privacy-audit.xlsx", canaries)
    screenshot_by_path = {str(item["path"]): item for item in evidence["artifacts"] if item["kind"] == "SCREENSHOT"}
    delivered_images = []
    for suffix in ("*.png", "*.jpg", "*.jpeg", "*.webp"):
        delivered_images.extend(path for path in directory.rglob(suffix) if path.is_file())
    for image in delivered_images:
        relative = image.resolve().relative_to(directory.resolve()).as_posix()
        artifact = screenshot_by_path.get(relative)
        if not artifact or not artifact.get("screenshot_review"):
            raise PrivacyError(f"Delivered image has no passed screenshot review: {relative}")
        extracted = extract_text(image)
        review_text(extracted, canaries)


def validate_delivery(directory: Path, canaries: list[str]) -> list[str]:
    missing = sorted(name for name in REQUIRED_FILES if not (directory / name).is_file())
    if missing:
        raise ContractError(f"Delivery is missing required files: {missing}")
    run = read_json(directory / "audit-run.json")
    validate_schema(run, "audit-run.schema.json", label="audit-run.json")
    if run["status"] != "COMPLETE":
        raise ContractError("Only a COMPLETE run can pass delivery validation")
    if run["technical_priority_rubric_version"] != PRIORITY_RUBRIC_VERSION:
        raise ContractError(f"This validator does not implement rubric {run['technical_priority_rubric_version']}")
    if run["finding_fingerprint_algo_version"] != FINGERPRINT_ALGO_VERSION:
        raise ContractError(f"This validator does not implement fingerprint algorithm {run['finding_fingerprint_algo_version']}")
    if run["identity_normalization_version"] != IDENTITY_NORMALIZATION_VERSION:
        raise ContractError(f"This validator does not implement identity normalization {run['identity_normalization_version']}")
    observations = list(iter_jsonl(directory / "observations.jsonl"))
    observation_ids: set[str] = set()
    observation_evidence: set[str] = set()
    for row in observations:
        validate_schema(row, "observation.schema.json", label=row["observation_id"])
        if row["run_id"] != run["run_id"]:
            raise ContractError(f"Observation run mismatch: {row['observation_id']}")
        if row["observation_id"] in observation_ids:
            raise ContractError(f"Duplicate observation ID: {row['observation_id']}")
        observation_ids.add(row["observation_id"])
        observation_evidence.add(row["evidence_id"])
    findings = load_findings(directory / "findings.json")
    fingerprints: set[str] = set()
    for item in findings:
        fingerprint = item["finding_fingerprint"]
        if fingerprint in fingerprints:
            raise ContractError(f"Duplicate canonical finding fingerprint: {fingerprint}")
        fingerprints.add(fingerprint)
        if finding_fingerprint(item["fingerprint_inputs"]) != fingerprint:
            raise ContractError(f"Finding fingerprint is not independently reproducible: {fingerprint}")
        if technical_priority(item["technical_priority_inputs"]) != item["technical_priority"]:
            raise ContractError(f"Technical priority is not independently reproducible: {fingerprint}")
        if item["technical_priority_rubric_version"] != run["technical_priority_rubric_version"]:
            raise ContractError(f"Finding/run priority rubric mismatch: {fingerprint}")
        if item["finding_fingerprint_algo_version"] != run["finding_fingerprint_algo_version"]:
            raise ContractError(f"Finding/run fingerprint algorithm mismatch: {fingerprint}")
        if item["root_cause_status"] == "CONFIRMED" and not item["confirmed_root_cause_evidence"]:
            raise ContractError(f"Confirmed root cause lacks confirmation evidence: {fingerprint}")
        if item["suspected_implementation_layer"] == "GTM_CONTAINER" and item["root_cause_status"] == "CONFIRMED" and not item["confirmed_root_cause_evidence"]:
            raise ContractError(f"GTM initiator was improperly promoted to confirmed root cause: {fingerprint}")
    declaration = read_json(directory / "declaration-diff.json")
    validate_schema(declaration, "declaration-diff.schema.json", label="declaration-diff.json")
    if declaration["run_id"] != run["run_id"]:
        raise ContractError("Declaration diff run identity mismatch")
    evidence = read_json(directory / "evidence-index.json")
    validate_schema(evidence, "evidence-index.schema.json", label="evidence-index.json")
    if evidence["run_id"] != run["run_id"]:
        raise ContractError("Evidence index run identity mismatch")
    evidence_ids: set[str] = set()
    for artifact in evidence["artifacts"]:
        evidence_id = artifact["evidence_id"]
        if evidence_id in evidence_ids:
            raise ContractError(f"Duplicate evidence ID: {evidence_id}")
        evidence_ids.add(evidence_id)
        path = (directory / artifact["path"]).resolve()
        if not path_within(path, directory):
            raise ContractError(f"Evidence path escapes delivery: {artifact['path']}")
        if not path.is_file() or sha256_file(path) != artifact["sha256"]:
            raise ContractError(f"Evidence file missing or hash mismatch: {artifact['path']}")
    for finding in findings:
        missing_evidence = set(finding["evidence_ids"]) - evidence_ids
        if missing_evidence:
            raise ContractError(f"Finding references missing evidence: {finding['finding_fingerprint']} -> {sorted(missing_evidence)}")
    source_status = {str(item["source_id"]): str(item["status"]) for item in run["source_checks"]}
    for check in run["source_checks"]:
        if check["status"] != "MATCHED" and not check["task_required"]:
            raise ContractError(f"Non-matching source lacks human re-verification task: {check['source_id']}")
    for finding in findings:
        if any(source_status.get(source_id) != "MATCHED" for source_id in finding["rule_source_ids"]):
            if finding["technical_test_status"] not in {"INCONCLUSIVE", "NOT_APPLICABLE", "NOT_TESTED"}:
                raise ContractError(f"Finding claims a conclusive rule result from an unverified source: {finding['finding_fingerprint']}")
    core = {"UNTOUCHED", "REJECTED", "ACCEPTED", "ACCEPTED_TO_WITHDRAWN", "PERSISTENCE_ACCEPTED", "PERSISTENCE_REJECTED"}
    state_gap_classes = {item["fingerprint_inputs"]["scenario_class"] for item in findings if item["finding_kind"] == "STATE_VERIFICATION_GAP"}
    for scenario in run["scenarios"]:
        if any(status == "REGISTERED" for status in scenario["capture_status"].values()):
            raise ContractError(f"Scenario retains a REGISTERED capture status: {scenario['scenario_id']}")
        if scenario["scenario_class"] in core and (scenario["status"] != "COMPLETE" or scenario["state_verified"] is not True):
            if scenario["scenario_class"] not in state_gap_classes:
                raise ContractError(f"Material scenario gap has no finding: {scenario['scenario_class']}")
    if run["scope"]["markets"] and not run["network_route"]["externally_verified_region"]:
        if any(item["test_context"]["market_claim_supported"] for item in findings):
            raise ContractError("Browser geolocation or locale was used to support an unverified market claim")
    remediation = read_json(directory / "remediation-handoff.json")
    validate_schema(remediation, "remediation-handoff.schema.json", label="remediation-handoff.json")
    if remediation["source_run_id"] != run["run_id"]:
        raise ContractError("Remediation handoff run identity mismatch")
    for item in remediation["handoffs"]:
        if item["source_finding_fingerprint"] not in fingerprints:
            raise ContractError("Remediation handoff references an unknown finding")
    recette_path = directory / "recette-handoff.json"
    gtm_fingerprints = {item["finding_fingerprint"] for item in findings if item["suspected_implementation_layer"] == "GTM_CONTAINER"}
    if recette_path.exists():
        recette = read_json(recette_path)
        validate_schema(recette, "recette-handoff.schema.json", label="recette-handoff.json")
        if recette["source_run_id"] != run["run_id"]:
            raise ContractError("Recette handoff run identity mismatch")
        for item in recette["handoffs"]:
            if item["source_finding_fingerprint"] not in gtm_fingerprints:
                raise ContractError("Recette handoff exists without a substantiated suspected-GTM finding")
    elif gtm_fingerprints:
        raise ContractError("A suspected-GTM finding lacks its supporting-only recette handoff")
    monitoring = read_json(directory / "monitoring-baseline.json")
    validate_schema(monitoring, "monitoring-baseline.schema.json", label="monitoring-baseline.json")
    if set(monitoring["finding_fingerprints"]) != fingerprints:
        raise ContractError("Monitoring baseline finding set differs from canonical findings")
    for output_name, output in run["outputs"].items():
        path = (directory / output["path"]).resolve()
        if not path_within(path, directory):
            raise ContractError(f"Output path escapes delivery: {output_name}")
        if not path.is_file():
            raise ContractError(f"Manifest output is missing: {output_name} -> {output['path']}")
        if output["sha256"] != sha256_file(path):
            raise ContractError(f"Manifest output hash mismatch: {output_name}")
    scan_delivery_files(directory, canaries, evidence)
    return [
        "schemas", "run-identities", "source-profile", "scenario-capture", "finding-fingerprint",
        "technical-priority", "evidence-cross-references", "handoff-authority", "privacy", "language", "workbook",
    ]


def main() -> int:
    args = parse_args()
    try:
        checks = validate_package()
        if not args.package_only:
            if not args.delivery:
                raise ContractError("delivery is required unless --package-only is used")
            canaries = load_restricted_canaries(args.canary_file)
            checks.extend(validate_delivery(args.delivery.resolve(), [item["value"] for item in canaries.values()]))
    except (ContractError, PrivacyError, OSError, ValueError) as exc:
        print(f"ERROR {exc}", file=sys.stderr)
        return 2
    print(f"PASS {len(checks)} validation checks")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
