from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from consent_runtime_core import (
    ContractError,
    assert_privacy_safe,
    iter_jsonl,
    load_profile,
    read_json,
    sha256_file,
    stable_id,
    utc_now,
    validate_schema,
    write_json,
    path_within,
)
from redact_and_index_evidence import build_index


PRIORITY_ORDER = ["URGENT", "HIGH", "MEDIUM", "LOW", "INFORMATIONAL"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a validated Markdown/XLSX consent audit delivery from canonical JSON.")
    parser.add_argument("delivery", type=Path, help="Directory containing audit-run.json, observations.jsonl, findings.json, and declaration-diff.json")
    return parser.parse_args()


def load_artifacts(directory: Path) -> tuple[dict[str, Any], list[dict[str, Any]], list[dict[str, Any]], dict[str, Any]]:
    run = read_json(directory / "audit-run.json")
    observations = list(iter_jsonl(directory / "observations.jsonl"))
    findings = read_json(directory / "findings.json")
    declaration = read_json(directory / "declaration-diff.json")
    validate_schema(run, "audit-run.schema.json", label="audit-run.json")
    if run["status"] != "COMPLETE":
        raise ContractError("Delivery can be built only for a COMPLETE run")
    if not isinstance(findings, list):
        raise ContractError("findings.json must be an array")
    for row in observations:
        validate_schema(row, "observation.schema.json", label=row["observation_id"])
        if row["run_id"] != run["run_id"]:
            raise ContractError("Observation run identity mismatch")
    for finding in findings:
        validate_schema(finding, "finding.schema.json", label=finding["finding_fingerprint"])
    validate_schema(declaration, "declaration-diff.schema.json", label="declaration-diff.json")
    if declaration["run_id"] != run["run_id"]:
        raise ContractError("Declaration diff run identity mismatch")
    return run, observations, findings, declaration


def build_handoffs(run: dict[str, Any], findings: list[dict[str, Any]], directory: Path) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    remediation = {
        "schema_version": "1.0.0",
        "source_run_id": run["run_id"],
        "generated_at": utc_now(),
        "handoffs": [],
    }
    recette = {
        "schema_version": "1.0.0",
        "source_run_id": run["run_id"],
        "generated_at": utc_now(),
        "handoffs": [],
    }
    for finding in findings:
        if finding["technical_test_status"] == "EXPECTED_BEHAVIOUR_OBSERVED":
            continue
        remediation["handoffs"].append(
            {
                "handoff_id": stable_id("REMED", run["run_id"], finding["finding_fingerprint"]),
                "source_finding_fingerprint": finding["finding_fingerprint"],
                "primary_owner": finding["primary_owner"],
                "contributing_owners": finding["contributing_owners"],
                "cause_status": finding["root_cause_status"],
                "requested_outcome": finding["proposed_outcome"],
                "prohibited_reinterpretations": [
                    "Do not treat this handoff as GTM, CMP, website, or policy mutation authority.",
                    "Do not convert this technical result into a legal-compliance conclusion.",
                ],
                "approval_state": "PROPOSED",
                "external_decisions": ["DPO or client applicability decision required"] if finding["legal_review_required"] else [],
                "retest_rule": "After an approved deployed change, rerun the exact clean browser scenario without GTM Preview and compare the stable slice.",
                "role": "MANUAL_ONLY",
            }
        )
        if finding["suspected_implementation_layer"] == "GTM_CONTAINER":
            recette["handoffs"].append(
                {
                    "handoff_id": stable_id("RECETTE", run["run_id"], finding["finding_fingerprint"]),
                    "source_finding_fingerprint": finding["finding_fingerprint"],
                    "source_evidence_ids": finding["evidence_ids"],
                    "suspected_gtm_container": None,
                    "reproduction": finding["retest_steps"],
                    "expected_behaviour": finding["expected_technical_behaviour"],
                    "observed_behaviour": finding["observed_fact"],
                    "consent_state": ", ".join(finding["scenario_classes"]),
                    "vendor_product_key": finding["fingerprint_inputs"]["vendor_product_key"],
                    "request_cues": sorted(set(finding["action_windows"] + [finding["fingerprint_inputs"]["location_pattern"]])),
                    "acceptance_rule": "Use Preview only to test the suspected GTM cause; preserve the browser finding and require a later Preview-independent browser rescan.",
                    "limitations": sorted(set(finding["limitations"] + ["GTM initiator evidence does not identify a tag, trigger, exception, or container object"])),
                    "role": "SUPPORTING_ONLY",
                    "verdict_authority": False,
                }
            )
    profile_freshness: dict[str, int] = {}
    for selection in run["rule_profiles"]:
        profile = load_profile(selection["profile_id"])
        for source in profile["sources"]:
            profile_freshness[str(source["source_id"])] = int(source["staleness_threshold_days"])
    monitoring = {
        "schema_version": "1.0.0",
        "source_run_id": run["run_id"],
        "created_at": utc_now(),
        "finding_fingerprint_algo_version": run["finding_fingerprint_algo_version"],
        "scenario_identities": sorted({str(item["scenario_class"]) for item in run["scenarios"]}),
        "finding_fingerprints": sorted({str(item["finding_fingerprint"]) for item in findings}),
        "allowed_comparison_fields": ["technical_test_status", "technical_priority", "rule_applicability", "attribution_confidence", "root_cause_status"],
        "freshness_requirements": profile_freshness,
        "role": "FUTURE_INTEGRATION_ONLY",
    }
    validate_schema(remediation, "remediation-handoff.schema.json", label="remediation handoff")
    validate_schema(recette, "recette-handoff.schema.json", label="recette handoff")
    validate_schema(monitoring, "monitoring-baseline.schema.json", label="monitoring baseline")
    for label, value in [("remediation handoff", remediation), ("recette handoff", recette), ("monitoring baseline", monitoring)]:
        assert_privacy_safe(value, label=label)
    write_json(directory / "remediation-handoff.json", remediation)
    if recette["handoffs"]:
        write_json(directory / "recette-handoff.json", recette)
    write_json(directory / "monitoring-baseline.json", monitoring)
    return remediation, recette, monitoring


def load_or_build_evidence(
    run: dict[str, Any], observations: list[dict[str, Any]], findings: list[dict[str, Any]], directory: Path
) -> dict[str, Any]:
    evidence_path = directory / "evidence-index.json"
    if not evidence_path.exists():
        return build_index(run, observations, findings, output=evidence_path, evidence_dir=directory / "evidence")
    evidence = read_json(evidence_path)
    validate_schema(evidence, "evidence-index.schema.json", label=str(evidence_path))
    if evidence["run_id"] != run["run_id"]:
        raise ContractError("Existing evidence index run identity does not match the delivery")
    required_ids = {str(row["evidence_id"]) for row in observations}
    required_ids.update(str(evidence_id) for finding in findings for evidence_id in finding["evidence_ids"])
    indexed_ids = {str(item["evidence_id"]) for item in evidence["artifacts"]}
    if required_ids - indexed_ids:
        raise ContractError(f"Existing evidence index is missing canonical evidence: {sorted(required_ids - indexed_ids)}")
    non_screenshot_ids = {str(item["evidence_id"]) for item in evidence["artifacts"] if item["kind"] != "SCREENSHOT"}
    if non_screenshot_ids != required_ids:
        raise ContractError("Existing evidence index contains stale non-screenshot evidence; rebuild it before delivery")
    for item in evidence["artifacts"]:
        path = (directory / item["path"]).resolve()
        if not path_within(path, directory) or not path.is_file() or sha256_file(path) != item["sha256"]:
            raise ContractError(f"Existing evidence artifact is missing or changed: {item['path']}")
    return evidence


def md_escape(value: Any) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")


def build_report(run: dict[str, Any], findings: list[dict[str, Any]], declaration: dict[str, Any], evidence: dict[str, Any], remediation: dict[str, Any]) -> str:
    outcomes = {
        "NO_SAMPLED_CONTRADICTION": "No sampled technical contradiction was detected under the selected and applicable tests.",
        "CONTRADICTIONS_OBSERVED": "One or more sampled technical contradictions were observed.",
        "MATERIAL_TESTS_INCONCLUSIVE": "Material tests remain inconclusive or untested.",
    }
    lines = [
        "# Consent Privacy Audit",
        "",
        outcomes.get(run["overall_technical_outcome"], "Material tests remain inconclusive or untested."),
        "",
        "This report is a bounded technical browser audit. It is not a legal opinion or a certification of GDPR, ePrivacy, CNIL, or whole-site compliance.",
        "",
        "## Scope and evidence boundary",
        "",
        f"- Run: `{md_escape(run['run_id'])}`",
        f"- Origin/environment: `{md_escape(run['site']['origin'])}` / `{md_escape(run['site']['environment'])}`",
        f"- Browser: Chromium `{md_escape(run['browser']['version'])}`; locale `{md_escape(run['browser']['locale'])}`; timezone `{md_escape(run['browser']['timezone'])}`",
        f"- Network route: `{md_escape(run['network_route']['route_id'])}`; verified region: `{md_escape(run['network_route']['externally_verified_region'] or 'not independently verified')}`",
        f"- Sampled URLs: {', '.join(f'`{md_escape(url)}`' for url in run['scope']['urls'])}",
        f"- Markets claimed: {', '.join(md_escape(item) for item in run['scope']['markets']) or 'none'}",
        f"- CMP adapter: `{md_escape(run['cmp']['adapter_id'] or 'unknown')}` at `{md_escape(run['cmp']['detection_confidence'])}` confidence",
        f"- Sanitized evidence artifacts: {len(evidence['artifacts'])}",
        "",
        "Authenticated areas, signed-in consent synchronization, multi-device behavior, GPC, unseen server forwarding, CMP back-office receipts, and whole privacy-notice review are outside v1.",
        "",
        "## Scenario results",
        "",
        "| Scenario | Status | State verified | Pages | Limitations |",
        "| --- | --- | --- | ---: | --- |",
    ]
    for scenario in run["scenarios"]:
        lines.append(f"| {md_escape(scenario['scenario_class'])} | {md_escape(scenario['status'])} | {md_escape(scenario['state_verified'])} | {len(scenario['pages'])} | {md_escape('; '.join(scenario['limitations']) or 'None recorded')} |")
    lines.extend(["", "## Findings", ""])
    for priority in PRIORITY_ORDER:
        rows = [item for item in findings if item["technical_priority"] == priority]
        if not rows:
            continue
        lines.extend([f"### {priority}", ""])
        for item in rows:
            lines.extend(
                [
                    f"#### {md_escape(item['title'])}",
                    "",
                    f"- Status/applicability: `{item['technical_test_status']}` / `{item['rule_applicability']}`",
                    f"- Rule: `{md_escape(item['rule_id'])}` ({md_escape(item['rule_authority'])})",
                    f"- Observed: {md_escape(item['observed_fact'])}",
                    f"- Expected technical behavior: {md_escape(item['expected_technical_behaviour'])}",
                    f"- Suspected layer/root cause: `{item['suspected_implementation_layer']}` / `{item['root_cause_status']}`",
                    f"- Owner: `{item['primary_owner']}`; contributors: {', '.join(f'`{owner}`' for owner in item['contributing_owners']) or 'none'}",
                    f"- Evidence: {', '.join(f'`{evidence_id}`' for evidence_id in item['evidence_ids'])}",
                    f"- Fingerprint: `{item['finding_fingerprint']}`",
                    "",
                ]
            )
    if not findings:
        lines.extend(["No canonical finding was generated for the sampled tests.", ""])
    lines.extend([
        "## Declaration consistency",
        "",
        "| Product | Direction | Status | Confidence |",
        "| --- | --- | --- | --- |",
    ])
    for item in declaration["items"]:
        lines.append(f"| {md_escape(item['display_name'])} | {item['direction']} | {item['status']} | {item['confidence']} |")
    if not declaration["items"]:
        lines.append("| No sampled tracker rows | — | NOT_VERIFIED | UNKNOWN |")
    lines.extend(["", "## Owner backlog", ""])
    by_owner: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_fp = {item["finding_fingerprint"]: item for item in findings}
    for handoff in remediation["handoffs"]:
        by_owner[handoff["primary_owner"]].append(handoff)
    for owner in sorted(by_owner):
        lines.extend([f"### {owner}", ""])
        for handoff in by_owner[owner]:
            finding = by_fp[handoff["source_finding_fingerprint"]]
            lines.append(f"- {md_escape(finding['title'])}: {md_escape(handoff['requested_outcome'])} Approval remains `{handoff['approval_state']}`.")
        lines.append("")
    lines.extend(["## Sources, limitations, and deferred work", ""])
    for check in run["source_checks"]:
        lines.append(f"- Source `{md_escape(check['source_id'])}`: `{check['status']}`; human task required: `{check['task_required']}`{'; ' + md_escape(check.get('note')) if check.get('note') else ''}")
    for exclusion in run["scope"]["exclusions"]:
        lines.append(f"- Excluded: {md_escape(exclusion)}")
    for boundary in run["scope"]["authenticated_boundaries"]:
        lines.append(f"- Authenticated boundary not covered in v1: {md_escape(boundary)}")
    lines.extend([
        "",
        "## Next step and retest",
        "",
        "Resolve proposed owner decisions and approvals. If a deployed change is made, rerun the exact clean browser scenario without GTM Preview. Use a supporting-only Preview handoff only where GTM is suspected, then preserve the independent production rescan as the final browser evidence.",
        "",
    ])
    return "\n".join(lines)


def excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True)
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text
    return text[:32767]


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([excel_value(value) for value in row])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    if sheet.max_row >= 2 and sheet.max_column >= 1:
        table_name = "T" + "".join(ch for ch in title if ch.isalnum())[:20]
        table = Table(displayName=table_name, ref=sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)
    for column in sheet.columns:
        width = min(60, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(run: dict[str, Any], findings: list[dict[str, Any]], declaration: dict[str, Any], evidence: dict[str, Any], remediation: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    counts = Counter(item["technical_priority"] for item in findings)
    add_sheet(workbook, "Summary", ["Metric", "Value"], [
        ("Run ID", run["run_id"]), ("Technical outcome", run["overall_technical_outcome"]),
        ("Origin", run["site"]["origin"]), ("Environment", run["site"]["environment"]),
        *[(f"Findings {priority}", counts.get(priority, 0)) for priority in PRIORITY_ORDER],
        ("No-legal-opinion boundary", "Bounded technical browser audit only"),
    ])
    add_sheet(workbook, "Scope & Coverage", ["Type", "Identity", "Status / detail"], [
        *[("URL", url, "Sampled") for url in run["scope"]["urls"]],
        *[("Market", market, f"Route region: {run['network_route']['externally_verified_region'] or 'unverified'}") for market in run["scope"]["markets"]],
        *[("Exclusion", value, "Excluded") for value in run["scope"]["exclusions"]],
        *[("Authenticated boundary", value, "Post-v1 / uncovered") for value in run["scope"]["authenticated_boundaries"]],
    ])
    add_sheet(workbook, "Scenario Results", ["Scenario ID", "Class", "Status", "State verified", "Context ID", "Pages", "Actions", "Limitations"], [
        (item["scenario_id"], item["scenario_class"], item["status"], item["state_verified"], item["context_id"], item["pages"], item["actions"], item["limitations"]) for item in run["scenarios"]
    ])
    add_sheet(workbook, "Findings", ["Priority", "Status", "Applicability", "Kind", "Title", "Rule", "Observed fact", "Expected", "Scenarios", "Actions", "Locations", "Layer", "Root cause", "Owner", "Contributors", "Evidence", "Fingerprint", "Legal review", "Limitations"], [
        (item["technical_priority"], item["technical_test_status"], item["rule_applicability"], item["finding_kind"], item["title"], item["rule_id"], item["observed_fact"], item["expected_technical_behaviour"], item["scenario_classes"], item["action_windows"], item["locations"], item["suspected_implementation_layer"], item["root_cause_status"], item["primary_owner"], item["contributing_owners"], item["evidence_ids"], item["finding_fingerprint"], item["legal_review_required"], item["limitations"]) for item in findings
    ])
    add_sheet(workbook, "Declaration Diff", ["Product key", "Display name", "Observed", "Declared", "Direction", "Status", "Confidence", "Declaration IDs", "Purpose comparison", "Duration comparison", "Legal review"], [
        (item["vendor_product_key"], item["display_name"], item["observed"], item["declared"], item["direction"], item["status"], item["confidence"], item["declaration_ids"], item["purpose_comparison"], item["duration_comparison"], item["legal_review_required"]) for item in declaration["items"]
    ])
    by_fp = {item["finding_fingerprint"]: item for item in findings}
    add_sheet(workbook, "Owner Backlog", ["Owner", "Contributors", "Priority", "Finding", "Requested outcome", "Cause status", "Approval", "External decisions", "Role"], [
        (item["primary_owner"], item["contributing_owners"], by_fp[item["source_finding_fingerprint"]]["technical_priority"], by_fp[item["source_finding_fingerprint"]]["title"], item["requested_outcome"], item["cause_status"], item["approval_state"], item["external_decisions"], item["role"]) for item in remediation["handoffs"]
    ])
    add_sheet(workbook, "Evidence Index", ["Evidence ID", "Kind", "Path", "SHA-256", "Sanitization", "Screenshot review"], [
        (item["evidence_id"], item["kind"], item["path"], item["sha256"], item["sanitization_status"], item["screenshot_review"]) for item in evidence["artifacts"]
    ])
    limitations = []
    for scenario in run["scenarios"]:
        limitations.extend((scenario["scenario_class"], item) for item in scenario["limitations"])
    limitations.extend(("Scope exclusion", item) for item in run["scope"]["exclusions"])
    limitations.extend(("Authenticated / post-v1", item) for item in run["scope"]["authenticated_boundaries"])
    limitations.extend(("Deferred", item) for item in ["GPC", "authenticated consent sync", "multi-device behavior", "unseen server forwarding", "CMP back-office receipts", "whole privacy-notice review"])
    add_sheet(workbook, "Limitations & Deferred Work", ["Area", "Limitation / deferred item"], limitations)
    workbook.properties.title = "Consent Privacy Audit"
    workbook.properties.subject = f"Technical browser audit {run['run_id']}"
    workbook.properties.creator = "consent-privacy-audit"
    workbook.save(path)


def main() -> int:
    args = parse_args()
    try:
        directory = args.delivery.resolve()
        run, observations, findings, declaration = load_artifacts(directory)
        evidence_path = directory / "evidence-index.json"
        evidence = load_or_build_evidence(run, observations, findings, directory)
        remediation, recette, monitoring = build_handoffs(run, findings, directory)
        report = build_report(run, findings, declaration, evidence, remediation)
        assert_privacy_safe(report, label="Markdown report")
        report_path = directory / "consent-privacy-audit.md"
        report_path.write_text(report, encoding="utf-8", newline="\n")
        workbook_path = directory / "consent-privacy-audit.xlsx"
        build_workbook(run, findings, declaration, evidence, remediation, workbook_path)
        outputs = {
            "evidence_index": evidence_path,
            "remediation_handoff": directory / "remediation-handoff.json",
            "monitoring_baseline": directory / "monitoring-baseline.json",
            "report": report_path,
            "workbook": workbook_path,
        }
        if recette["handoffs"]:
            outputs["recette_handoff"] = directory / "recette-handoff.json"
        for name, path in outputs.items():
            run["outputs"][name] = {"path": path.name, "sha256": sha256_file(path)}
        validate_schema(run, "audit-run.schema.json", label="audit-run.json")
        write_json(directory / "audit-run.json", run)
    except (ContractError, OSError, ValueError) as exc:
        print(f"ERROR {exc}", file=sys.stderr)
        return 2
    print(directory)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
